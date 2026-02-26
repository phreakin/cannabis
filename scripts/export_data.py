#!/usr/bin/env python3
"""
CLI export script: exports collected data to CSV, JSON, GeoJSON, or XLSX.

Usage:
    python scripts/export_data.py --format csv
    python scripts/export_data.py --format geojson --state CO
    python scripts/export_data.py --format csv --state WA --category dispensary
    python scripts/export_data.py --format json --output my_data.json
    python scripts/export_data.py --format xlsx --output cannabis_export.xlsx
    python scripts/export_data.py --format csv --limit 10000
"""
import argparse
import csv
import json
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv()


def build_query(session, state=None, category=None, source_id=None, has_gps=False,
                limit=None):
    from src.storage.models import RawRecord

    q = session.query(RawRecord)
    if state:
        q = q.filter(RawRecord.state == state.upper())
    if category:
        q = q.filter(RawRecord.category.ilike(f"%{category}%"))
    if source_id:
        q = q.filter(RawRecord.source_id == int(source_id))
    if has_gps:
        q = q.filter(
            RawRecord.latitude.isnot(None),
            RawRecord.longitude.isnot(None),
        )
    q = q.order_by(RawRecord.state, RawRecord.category, RawRecord.name)
    if limit:
        q = q.limit(limit)
    return q


STANDARD_FIELDS = [
    "id", "source_id", "run_id", "state", "category", "subcategory",
    "name", "license_number", "license_type", "license_status",
    "address", "city", "zip_code", "county",
    "latitude", "longitude", "phone", "email", "website",
    "record_date", "license_date", "expiry_date", "collected_at",
]


def record_to_dict(record) -> dict:
    return {f: getattr(record, f, None) for f in STANDARD_FIELDS}


def export_csv(records, output_path: str):
    print(f"Exporting CSV to {output_path}...")
    count = 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = None
        for record in records:
            row = record_to_dict(record)
            if writer is None:
                writer = csv.DictWriter(f, fieldnames=list(row.keys()))
                writer.writeheader()
            # Serialize datetimes
            for k, v in row.items():
                if hasattr(v, "isoformat"):
                    row[k] = v.isoformat()
            writer.writerow(row)
            count += 1
            if count % 10000 == 0:
                print(f"  {count:,} records written...")
    print(f"[OK] Exported {count:,} records")
    return count


def export_json(records, output_path: str):
    print(f"Exporting JSON to {output_path}...")
    data = []
    count = 0
    for record in records:
        row = record_to_dict(record)
        for k, v in row.items():
            if hasattr(v, "isoformat"):
                row[k] = v.isoformat()
        # Include raw data if available
        if hasattr(record, "record_data") and record.record_data:
            row["_raw"] = record.record_data
        data.append(row)
        count += 1

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str, ensure_ascii=False)
    print(f"[OK] Exported {count:,} records")
    return count


def export_geojson(records, output_path: str):
    print(f"Exporting GeoJSON to {output_path}...")
    features = []
    skipped = 0
    count = 0
    for record in records:
        if not record.latitude or not record.longitude:
            skipped += 1
            continue
        props = record_to_dict(record)
        # Remove coordinates from properties (they're in geometry)
        lat = props.pop("latitude")
        lng = props.pop("longitude")
        for k, v in props.items():
            if hasattr(v, "isoformat"):
                props[k] = v.isoformat()
        features.append({
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [float(lng), float(lat)],
            },
            "properties": props,
        })
        count += 1

    fc = {
        "type": "FeatureCollection",
        "properties": {
            "exported_at": datetime.utcnow().isoformat(),
            "total_features": count,
        },
        "features": features,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(fc, f, indent=2, default=str, ensure_ascii=False)

    if skipped:
        print(f"  (Skipped {skipped:,} records without GPS coordinates)")
    print(f"[OK] Exported {count:,} geo features")
    return count


def export_xlsx(records, output_path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[ERR] openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    print(f"Exporting Excel to {output_path}...")

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Records"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1a5c2a")

    headers = STANDARD_FIELDS
    for col, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Track by category for separate sheets
    by_category = {}
    count = 0

    for row_idx, record in enumerate(records, 2):
        row = record_to_dict(record)
        for col, field in enumerate(headers, 1):
            val = row.get(field)
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            ws_all.cell(row=row_idx, column=col, value=val)

        cat = record.category or "Uncategorized"
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(row)
        count += 1

    # Category sheets
    for cat, rows in sorted(by_category.items()):
        safe_name = cat[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=safe_name)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, row in enumerate(rows, 2):
            for col, field in enumerate(headers, 1):
                val = row.get(field)
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                ws.cell(row=row_idx, column=col, value=val)

    wb.save(output_path)
    print(f"[OK] Exported {count:,} records across {len(by_category)+1} sheets")
    return count


def main():
    parser = argparse.ArgumentParser(
        description="Export cannabis data to CSV, JSON, GeoJSON, or XLSX"
    )
    parser.add_argument("--format", "-f", choices=["csv", "json", "geojson", "xlsx"],
                        default="csv", help="Output format (default: csv)")
    parser.add_argument("--output", "-o", default=None,
                        help="Output file path (default: auto-generated)")
    parser.add_argument("--state", help="Filter by state abbreviation (e.g. CO)")
    parser.add_argument("--category", help="Filter by category (partial match)")
    parser.add_argument("--source", help="Filter by source ID (database integer ID)")
    parser.add_argument("--gps-only", action="store_true",
                        help="Only export records with GPS coordinates")
    parser.add_argument("--limit", type=int, default=None,
                        help="Maximum records to export")
    parser.add_argument("--db-url", default=None)
    args = parser.parse_args()

    db_url = args.db_url or os.environ.get(
        "DATABASE_URL", "sqlite:///data/cannabis_aggregator.db"
    )

    from src.storage.database import init_db, session_scope
    if db_url.startswith("sqlite:///"):
        db_path = db_url.replace("sqlite:///", "")
        os.makedirs(os.path.dirname(os.path.abspath(db_path)), exist_ok=True)
    init_db(db_url)

    # Determine output path
    if not args.output:
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        parts = ["cannabis_data"]
        if args.state:
            parts.append(args.state.lower())
        if args.category:
            parts.append(args.category.lower().replace(" ", "_"))
        parts.append(timestamp)
        args.output = os.path.join(
            "data", "processed",
            "_".join(parts) + f".{args.format}"
        )

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)

    print(f"Database:  {db_url}")
    print(f"Format:    {args.format}")
    print(f"Output:    {args.output}")
    if args.state:    print(f"State:     {args.state}")
    if args.category: print(f"Category:  {args.category}")
    if args.gps_only: print(f"GPS only:  yes")
    if args.limit:    print(f"Limit:     {args.limit:,}")
    print()

    with session_scope() as session:
        q = build_query(
            session,
            state=args.state,
            category=args.category,
            source_id=args.source,
            has_gps=args.gps_only or args.format == "geojson",
            limit=args.limit,
        )

        # Count first
        total = q.count()
        print(f"Records matched: {total:,}")
        if total == 0:
            print("No records to export.")
            return

        records = q.all()

    fmt = args.format
    if fmt == "csv":
        export_csv(iter(records), args.output)
    elif fmt == "json":
        export_json(iter(records), args.output)
    elif fmt == "geojson":
        export_geojson(iter(records), args.output)
    elif fmt == "xlsx":
        export_xlsx(iter(records), args.output)

    print(f"\nOutput file: {os.path.abspath(args.output)}")
    size = os.path.getsize(args.output)
    print(f"File size:   {size / 1024:.1f} KB" if size < 1_000_000
          else f"File size:   {size / 1_000_000:.1f} MB")


if __name__ == "__main__":
    main()
